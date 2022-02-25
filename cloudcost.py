# PSQL stuff
import errno
import psycopg2
import psycopg2.extras
from psycopg2 import sql

# Python Standard stuff
import importlib
from importlib import util
import argparse
import inspect
from datetime import datetime
from getpass import getpass
import configparser
import requests
from os import path
import json
import os
from pathlib import Path
import subprocess

# XLSX support
from openpyxl import Workbook

# Upload file to mattermost
def upload_file(file, server, channel_id, token, failed):
    # Simple auth header
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type':'application/json'
    }

    # Payload for initiating the file upload
    payload = {
        'channel_id': channel_id,
        'filename': path.basename(file)
    }

    # File description
    files = {
        'file': open(file, 'rb'),
        'filename': path.basename(file)
    }

    # Create a new upload and grab the ID
    print("Uploading...")
    x = requests.post(f'{server}/api/v4/files', headers=headers, params=payload, data=open(file, 'rb'))
    js = json.loads(x.text)
    if not x.ok:
        raise Exception(f'Failed to upload {file}. Response:\n{json.dumps(js,indent=4)}')
    upload_id = js['file_infos'][0]['id']

    # Payload for post, attach the file we just uploaded
    payload = {
            'channel_id': channel_id,
            'message': '\n### Today\'s cloud cost!',
            'file_ids': [upload_id]
        }

    # Post the file
    print("Done.")
    x = requests.post(f'{server}/api/v4/posts', headers=headers, json=payload)
    if not x.ok:
        raise Exception(f'Failed to post message. Response:\n{json.dumps(x.json(),indent=4)}')
        
def post_machines(providers, server, channel_id, token):
    # Simple auth header
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type':'application/json'
    }
    
    # Create our attachments for the message, MM has slack style attachments
    attachments = []
    for iaas, vms in providers.items():
        for vm, info in vms.items():
            attachment = {
                'fallback': f"{vm} was billed for {info['hours']} hours in {iaas}",
                'title': 'You have been billed for a VM alive longer than 7 days',
                'color': '#FF0000',
                'fields': [
                    {'short': True, 'title': 'VM Name', 'value': vm},
                    {'short': True, 'title': '', 'value': ''},
                    {'short': True, 'title': 'Provider', 'value': iaas},
                    {'short': True, 'title': 'Account', 'value': info['account']},
                    {'short': True, 'title': 'Bill/Invoice', 'value': info['bill']},
                    {'short': True, 'title': 'Hours Billed', 'value': f"{round(float(info['hours']), 2):.2f}"},
                ]
            }
            attachments.append(attachment)
    
    # Set up our message payload & send
    payload = {
        'channel_id': channel_id,
        'message': '# Machines billed for > 7 days',
        'props': {
            'attachments': attachments
        }
    }
    
    x = requests.post(f'{server}/api/v4/posts', headers=headers, json=payload)
    if not x.ok:
        raise Exception(f'Failed to post message. Response:\n{json.dumps(x.json(),indent=4)}')
        
def post_failures(failures, server, channel_id, token):
    # Auth headers
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type':'application/json'
    }
    
    # Create our attachments for the message, MM has slack style attachments
    attachments = []
    for fail in failures:
        attachment = {
            'fallback': f"{fail['name']} in {fail['iaas']} failed",
            'title': 'An account has failed to report',
            'color': '#FF0000',
            'fields': [
                {'short': True, 'title': 'Provider', 'value': fail['iaas']},
                {'short': True, 'title': 'Account', 'value': fail['name']},
                {'short': False, 'title': 'Error', 'value': f"```\n{fail['error']}\n```"},
            ]
        }
        attachments.append(attachment)
    
    # Set up our message payload & send
    payload = {
        'channel_id': channel_id,
        'message': '# Failures',
        'props': {
            'attachments': attachments
        }
    }
    
    x = requests.post(f'{server}/api/v4/posts', headers=headers, json=payload)
    if not x.ok:
        raise Exception(f'Failed to post message. Response:\n{json.dumps(x.json(),indent=4)}')

# Quick function to retry something 5 times
def retry(func, *args, **kwargs):
    for i in range(5):
        try:
            func(*args, **kwargs)
        except Exception as err:
            # Failed but we have retries left
            if i < 4:
                # Print and continue
                print(err)
                continue
            else:
                # Retries exceeded, pass the exception up
                raise
        else:
            # Success, break from retry loop
            break

# Creates some headers for the excel sheet
# Returns the workbook, worksheet and starting row
def create_xlsx():
    # Workbook object
    wb = Workbook()

    # Grab active sheet, change title
    ws = wb.active
    ws.title = "Cloud Cost"

    # Reporting dates
    ws['A1'] = 'Report Made:'
    ws['B1'] = datetime.today()
    ws['B1'].number_format = 'yyyy-mm-dd'

    # Headers
    ws['A3'] = 'Provider'
    ws['B3'] = 'Billing Start'
    ws['C3'] = 'Billing End'
    ws['D3'] = 'Account Name'
    ws['E3'] = 'Current Invoice'
    ws['F3'] = 'Balance'

    # Return workbook, worksheet, starting row index
    return wb, ws, 4

# Queries DB and runs cost against all accounts, all providers
def run_cost(cur, **kwargs):
    args = kwargs['args']
    conf = kwargs['conf']
    
    if args.nopost:
        print('Not Posting.')
    
    # Set up our workbook
    wb, ws, row = create_xlsx()

    query = sql.SQL('select iaas, name, cred, enable from get_accounts({iaas});').format(
            iaas = sql.Literal([args.iaas] if 'iaas' in args else None)
        )
    cur.execute(query)

    # fetchall() will return us a dictionary of lists
    accounts = cur.fetchall()

    # Now we loop through each provider
    failed = []
    vms = {}
    for account in accounts:
        provider = account['iaas']
        name = account['name']
        
        # Skip if this isn't the account we want
        if 'account' in args and args.account != name:
            continue
        
        # Skip if account disabled
        if not account['enable']:
            print(f'Skipping account {provider} - {name}')
            continue
    
        try:
            # Import a module named with the provider from the providers directory
            module = importlib.import_module("providers.{}".format(provider))

            # Call out to provider module's cost() function
            costs = module.cost(name, **account['cred'])
            
            # Wrap in another try block so we don't kill cost if this fails
            try:
                # Only run if this provider has this implemented and start of new billing cycle
                today = datetime.utcnow().isoformat()
                if hasattr(module, 'life') and cost.startDate[:10] == today[:10]:
                    pvms = module.life(name, **account['cred'])
                
                    if len(pvms) > 0:
                        vms[account] = pvms
            except Exception as err:
                print(f"Failed to run life() on {provider} {name}: {err}")

            # Spill to excel
            # Since we now return a list of CostItems to accomodate for
            # Bluemix and Softlayer returning both previous and current billing
            # periods, loop through the list of CostItems returned
            # TODO: Gotta be a neater way to do this
            for cost in costs:
                ws[f'A{row}'] = provider
                # Split the string, if it contains more than a date, we only want the date
                ws[f'B{row}'] = cost.startDate[:10] if cost.startDate else None
                ws[f'C{row}'] = cost.endDate[:10] if cost.endDate else None
                ws[f'D{row}'] = name
                ws[f'E{row}'] = f"{round(float(cost.cost), 2):.2f}"
                ws[f'F{row}'] = cost.balance
                row = row + 1

                print("{} total cost to month is {}".format(name, cost))


        # Two things can lead here, module import failing and sql query failure
        # Skip this provider in this case
        except BaseException as err:
            print(f"{provider} {name} Failed with {err}")
            failed.append({'iaas': provider, 'name': name, 'error': err})

    # Save to the workbook
    fname = "/tmp/cloudcost{}.xlsx".format(datetime.today().strftime("%Y-%m-%d"))
    wb.save(fname)

    if args.nopost:
        return

    # Try posting our failed accounts
    if failed:
        # Retry upload 5 times
        retry(post_failures, failed, **conf['mattermost'])
        
    # Try posting our VM lifetimes
    if vms:
        # Retry upload 5 times
        retry(post_machines, vms, **conf['mattermost'])
            
    # Retry upload 5 times
    retry(upload_file, fname, **conf['mattermost'], failed=failed)
        
def run_life(cur, **kwargs):
    args = kwargs['args']
    provider = args.iaas
    account_name = args.account
    conf = kwargs['conf']
    
    # Grab all of our accounts optionally filter by provider name
    cur.execute(sql.SQL('select iaas, name, cred from get_accounts({iaas});').format(
        iaas = sql.Literal(provider),
    ))
    accounts = cur.fetchall()
    
    vms = {}
    for account in accounts:
        iaas = account['iaas']
        name = account['name']
        # Didn't add a specific account filter to the SQL
        if account_name is not None and name != account_name:
            continue
        
        try:
            # Import module, try the life commmand
            module = importlib.import_module("providers.{}".format(iaas))
            
            if hasattr(module, 'life'):
                print(f'Checking {name}...')
                pvms = module.life(name, **account['cred'])
                
                if pvms:
                    vms[iaas] = pvms

        except Exception as err:
            # Don't really do anything other than print a message
            print(f'Failed to run machine billing lifespan for provider {iaas} {err}')
    if args.nopost:
        print('Not Posting.')
        print(json.dumps(vms))
    elif vms:
        retry(post_machines, vms, **conf['mattermost'])

# Add a new account to the DB
def add_account(cur, **kwargs):
    args = kwargs['args']
    provider = args.iaas
    account = args.account
    # Check if the iaas already exists in the DB
    # if so just add it, if not try to locate its module and add a new provider
    
    cur.execute(sql.SQL('select get_iaas_id({iaas});').format(
        iaas = sql.Literal(provider)
    ))
    
    if cur.fetchone()[0] is None:
        try:
            module = importlib.import_module("providers.{}".format(provider))
        except:
            raise Exception(f'The provider {provider} does not have an implementation')
        cur.execute(sql.SQL('select create_iaas({iaas});').format(
            iaas = sql.Literal(provider)
        ))
    
    try:
        module = importlib.import_module("providers.{}".format(provider))

        # Get argument names from module's run command
        argspec = inspect.getfullargspec(module.cost)
        # argspec[0] is a list of names of standard arguments
        cols = argspec[0]
        
        cred = dict((f'{a}', getpass(f'{a}: ')) for a in filter(lambda a: a != 'account_name',cols))

        # Build arbritary insert using the psycopg2 sql extension
        # Need to convert everything into Identifier and Literal for this to work so map cols and vals
        insert = sql.SQL("select create_account({iaas}, {name}, {cred}, true);").format(
                iaas = sql.Literal(provider),
                name = sql.Literal(account),
                cred = sql.Literal(psycopg2.extras.Json(cred))
            )

        # Execute query and commit change
        cur.execute(insert)
        cur.connection.commit()

    # For now die here, need to find what exceptions can be thrown here
    except BaseException as err:
            print(f"Unexpected {err=}, {type(err)=}")
            raise

# Update an account
def update_account(cur, **kwargs):
    args = kwargs['args']
    provider = args.iaas
    account = args.account

    # First check if provider module exists
    # It doesn't we can't do anything
    try:
        module = importlib.import_module("providers.{}".format(provider))

        # Get argument names from module's run command
        argspec = inspect.getfullargspec(module.cost)
        # argspec[0] is a list of names of standard arguments
        # In this case just remove account_name from the list since we won't touch it
        cols = filter(lambda a: a != 'account_name', argspec[0])
        
        cred = dict((f'{a}', getpass(f'{a}: ')) for a in filter(lambda a: a != 'account_name',cols))

        # Build arbritary insert using the psycopg2 sql extension
        # Need to convert everything into Identifier and Literal for this to work so map cols and vals
        insert = sql.SQL("select create_account({iaas}, {name}, {cred}, false);").format(
                iaas = sql.Literal(provider),
                name = sql.Literal(account),
                cred = sql.Literal(psycopg2.extras.Json(cred))
            )
        
        cur.execute(insert)

    # For now die here, need to find what exceptions can be thrown here
    except BaseException as err:
        # Rollback here since we failed, pass exception up
        cur.connection.rollback()
        print(f"Unexpected {err=}, {type(err)=}")
        raise

    else:
        # Commit the changes only if we succeeded the try block
        cur.connection.commit()

# List either providers available or accounts (optionally in provider)
def list_account(cur, **kwargs):
    args = kwargs['args']
    # List all providers implemented
    if args.type == 'providers':
        cur.execute('select get_iaas();')
        for iaas in cur.fetchone()[0]:
            print(iaas)
    # List of all accounts
    # Takes an optional --iaas argument, list of providers to list from
    elif args.type == 'accounts':
        # Grab a list of all tables in the DB (should be a list of provider names)
        # this is postgresql specific
        query = sql.SQL('select iaas, name from get_accounts({iaas});').format(
            iaas = sql.Literal(args.iaas)
        )
        cur.execute(query)

        # fetchall() will return us a dictionary of lists
        # we only want the first value of each list
        accounts = cur.fetchall()
        
        # Just loop through every row in every table as filtered above and print account_name
        for account in accounts:
            print(f"{account['iaas']}: {account['name']}")

# Remove an account from the DB
def remove_account(cur, **kwargs):
    args = kwargs['args']
    provider = args.iaas
    account = args.account
    # Wrap in a try block, no exceptions should ever be thrown here but in case
    try:
        # Build our query
        query = sql.SQL("select delete_account({provider},{account});").format(
            provider = sql.Literal(provider),
            account = sql.Literal(account)
        )
        # Run, check # of rows deleted to check for success
        cur.execute(query)
        print(f"Success: {cur.statusmessage}")
        cur.connection.commit()
    except Exception as err:
        print(f'{err=}, {type(err)=}')
        
def order_accounts(cur, **kwargs):
    args = kwargs['args']
    
    cur.execute('select iaas, name from get_accounts();')
    accounts = cur.fetchall()
    
    with open('/tmp/order.lst', 'w') as f:
        for account in accounts:
            f.write(f"{account['iaas']} | {account['name']}\n")
    
    cmd = os.environ.get('EDITOR', 'vi') + ' /tmp/order.lst'
    subprocess.call(cmd, shell=True)
    
    with open('/tmp/order.lst', 'r') as f:
        for i, line in enumerate(f.readlines()):
            (iaas,name) = line.strip().replace(' ','').split('|')
            cur.execute(sql.SQL('select set_order({iaas},{name},{order});').format(
                iaas = sql.Literal(iaas),
                name = sql.Literal(name),
                order = sql.Literal(i)
            ))
            
    print('Order set.')
    cur.connection.commit()

def main(args):
    # Connect to our postgres database
    conn = None
    try:
        # Read in our config file
        conf = configparser.ConfigParser()
        conf.read("/etc/cloudcost/cloudcost.conf")

        # Directly pass the config file as argument
        # ** formats it to pass the dictionary as named arguments
        conn = psycopg2.connect(**conf['database'])

        # Create a cursor, use this to interact with the DB
        # Cursor has support for "with" intrinsic (autocleanup outside of scope)
        # Read out results as a dictionary object
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            
            # Fork here depending on if we're running cost or adding a provider
            # Handle this by simpling calling the set function by argparse
            # We can pass anything the functions may need since they're declared
            # as kwargs
            args.func(cur, args=args, conf=conf)
                

    # Did we screw up connecting to database?
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)

    # Some other non-recoverable exception
    except BaseException as err:
        print(f"Unexpected {err=}, {type(err)=}")

    finally:
        # Need to clean up connection if made it
        if conn is not None:
            conn.close()

# Python standard, do nothing if imported
if __name__ == '__main__':
    # Parse our arguments
    parser = argparse.ArgumentParser()
    parser.set_defaults(func=run_cost)
    parser.add_argument('--nopost', help='do not post to MM', action='store_true')
    subparsers = parser.add_subparsers(help='sub-commands, type <command> --help to get more information')
    
    sub_cost = subparsers.add_parser('cost', help='runs the cost function and posts to MM')
    sub_cost.add_argument('--iaas', type=str, required=False, help='iaas to modify account in')
    sub_cost.add_argument('--account', type=str, required=False, help='account to modify')
    sub_cost.set_defaults(func=run_cost)
    
    sub_life = subparsers.add_parser('life', help='runs a check on the previous invoice and alerts for things alive longer than a time')
    sub_life.add_argument('--iaas', action='extend', nargs='+', type=str, required=False, help='optional list of providers to list accounts from')
    sub_life.add_argument('--account', type=str, required=False, help='single account to run this against')
    sub_life.set_defaults(iaas=None, func=run_life)
    
    sub_list = subparsers.add_parser('list', help='list accounts or providers')
    sub_list.add_argument('type', choices=['accounts', 'providers'], help='the item to list')
    sub_list.add_argument('--iaas', action='extend', nargs='+', type=str, required=False, help='optional list of providers to list accounts from')
    sub_list.set_defaults(iaas=None, func=list_account)
    
    sub_update = subparsers.add_parser('update', help='update an exist account\'s credentials')
    sub_update.add_argument('--iaas', type=str, required=True, help='iaas to modify account in')
    sub_update.add_argument('--account', type=str, required=True, help='account to modify')
    sub_update.set_defaults(func=update_account)
    
    sub_add = subparsers.add_parser('add', help='add a new account')
    sub_add.add_argument('--iaas', type=str, required=True, help='iaas to add account to')
    sub_add.add_argument('--account', type=str, required=True, help='account name to add')
    sub_add.set_defaults(func=add_account)
    
    sub_remove = subparsers.add_parser('remove', help='remove an account')
    sub_remove.add_argument('--iaas', type=str, required=True, help='iaas to remove account from')
    sub_remove.add_argument('--account', type=str, required=True, help='account name to remove')
    sub_remove.set_defaults(func=remove_account)
    
    sub_order = subparsers.add_parser('order', help='set the ordering of accounts')
    sub_order.set_defaults(func=order_accounts)

    args = parser.parse_args()

    main(args)
